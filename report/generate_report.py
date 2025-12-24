import psycopg2
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import pandas as pd

# Database connection parameters
db_params = {
    'dbname': 'your_database',
    'user': 'your_username',
    'password': 'your_password',
    'host': 'localhost',
    'port': '5432'
}

# Function to convert duration in days to HH:MM:SS
def days_to_hhmmss(days):
    if not isinstance(days, (int, float)) or pd.isna(days):
        return "00:00:00"
    seconds = days * 86400  # Convert days to seconds
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds = int(seconds % 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

# Function to fetch data from PostgreSQL
def fetch_machine_data(date, shift):
    try:
        conn = psycopg2.connect(**db_params)
        cursor = conn.cursor()
        
        # Query machine state data for the given date and shift
        query = """
        SELECT time_start, time_end, status, status_code, message, duration
        FROM machine_state
        WHERE date_trunc('day', time_start) = %s AND shift = %s
        ORDER BY time_start
        """
        cursor.execute(query, (date, shift))
        data = cursor.fetchall()
        
        # Convert to DataFrame for easier processing
        columns = ['time_start', 'time_end', 'status', 'status_code', 'message', 'duration']
        df = pd.DataFrame(data, columns=columns)
        
        cursor.close()
        conn.close()
        return df
    except Exception as e:
        print(f"Error fetching data: {e}")
        return None

# Function to process data for shift and stoppage analysis
def process_data(df):
    # Calculate total running duration (status = 1)
    running_duration = df[df['status'] == 1]['duration'].sum()
    
    # Calculate total stoppage duration (status = 2 or 4)
    stoppage_duration = df[df['status'].isin([2, 4])]['duration'].sum()
    
    # Stoppage analysis: group by message to get total duration and frequency
    stoppage_df = df[df['status'].isin([2, 4])][['message', 'duration']].groupby('message').agg(
        total_stoppage_duration=('duration', 'sum'),
        frequency=('duration', 'count')
    ).reset_index()
    
    # Find highest duration stoppage reason
    highest_duration = stoppage_df.loc[stoppage_df['total_stoppage_duration'].idxmax()]
    highest_duration_reason = highest_duration['message']
    highest_duration_value = highest_duration['total_stoppage_duration']
    
    # Find most frequent stoppage reason
    most_frequent = stoppage_df.loc[stoppage_df['frequency'].idxmax()]
    most_frequent_reason = most_frequent['message']
    most_frequent_count = most_frequent['frequency']
    
    return {
        'running_duration': running_duration,
        'stoppage_duration': stoppage_duration,
        'highest_duration_reason': highest_duration_reason,
        'highest_duration_value': highest_duration_value,
        'most_frequent_reason': most_frequent_reason,
        'most_frequent_count': most_frequent_count,
        'stoppage_analysis': stoppage_df
    }

# Function to create Excel report
def create_excel_report(df, analysis, date, shift, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Machine Operations Report"

    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Report Header
    ws['A1'] = "Machine Operations Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Date: {date}"
    ws['A3'] = f"Shift: {shift}"
    
    # Machine State Table
    ws['A5'] = "Machine state table"
    ws['A5'].font = header_font
    
    headers = ['time_start', 'time_end', 'status', 'status_code', 'message', 'duration']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # Write machine state data
    for row_idx, row in df.iterrows():
        for col_idx, col in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx + 7, column=col_idx)
            value = row[col]
            if col in ['time_start', 'time_end'] and pd.notna(value):
                value = value.strftime('%Y-%m-%d %H:%M:%S%z')
            elif col == 'duration' and pd.notna(value):
                value = days_to_hhmmss(value)
            cell.value = value
            cell.border = border
            cell.alignment = center_align if col != 'message' else left_align
    
    # Shift Analysis
    row_start = len(df) + 9
    ws[f'A{row_start}'] = "Shift analysis"
    ws[f'A{row_start}'].font = header_font
    
    shift_headers = ['Parameter', 'Output (HH:MM:SS)']
    for col, header in enumerate(shift_headers, start=1):
        cell = ws.cell(row=row_start + 1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    shift_data = [
        ('Total running duration', days_to_hhmmss(analysis['running_duration'])),
        ('Total stoppage duration', days_to_hhmmss(analysis['stoppage_duration'])),
        ('Highest duration stoppage reason', f"{analysis['highest_duration_reason']} ({days_to_hhmmss(analysis['highest_duration_value'])})"),
        ('Most frequent stoppage reason', f"{analysis['most_frequent_reason']} ({analysis['most_frequent_count']} times)")
    ]
    
    for idx, (param, value) in enumerate(shift_data, start=row_start + 2):
        ws[f'A{idx}'].value = param
        ws[f'A{idx}'].border = border
        ws[f'A{idx}'].alignment = left_align
        ws[f'B{idx}'].value = value
        ws[f'B{idx}'].border = border
        ws[f'B{idx}'].alignment = center_align
    
    # Stoppage Analysis
    row_start += 6
    ws[f'A{row_start}'] = "Stoppage analysis"
    ws[f'A{row_start}'].font = header_font
    
    stoppage_headers = ['message', 'total_stoppage_duration', 'frequency']
    for col, header in enumerate(stoppage_headers, start=1):
        cell = ws.cell(row=row_start + 1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    for idx, row in analysis['stoppage_analysis'].iterrows():
        ws[f'A{row_start + 2 + idx}'].value = row['message']
        ws[f'A{row_start + 2 + idx}'].border = border
        ws[f'A{row_start + 2 + idx}'].alignment = left_align
        ws[f'B{row_start + 2 + idx}'].value = days_to_hhmmss(row['total_stoppage_duration'])
        ws[f'B{row_start + 2 + idx}'].border = border
        ws[f'B{row_start + 2 + idx}'].alignment = center_align
        ws[f'C{row_start + 2 + idx}'].value = row['frequency']
        ws[f'C{row_start + 2 + idx}'].border = border
        ws[f'C{row_start + 2 + idx}'].alignment = center_align
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(filename)
    print(f"Report saved as {filename}")

# Main execution
if __name__ == "__main__":
    date = "2025-05-23"
    shift = "A"
    filename = "Report_Machine17_16May2025_ShiftA.xlsx"
    
    # Fetch data
    df = fetch_machine_data(date, shift)
    if df is not None:
        # Process data
        analysis = process_data(df)
        # Create Excel report
        create_excel_report(df, analysis, date, shift, filename)
